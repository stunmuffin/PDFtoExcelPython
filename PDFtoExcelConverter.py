# ##### BEGIN GPL LICENSE BLOCK #####
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.
# ##### END GPL LICENSE BLOCK #####
#made by stunmuffin(KB)


import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pandas as pd
import os

file_pdf_name = "FileNameHere"
pdf_input_file_path = file_pdf_name + ".pdf"  
pdf_output_file_path_xlsx = file_pdf_name + ".xlsx"

# Initialize an empty DataFrame to store the combined tables
combined_df = pd.DataFrame()

with pdfplumber.open(pdf_input_file_path) as pdf:
    # Loop through all pages
    for page_num in range(len(pdf.pages)):
        current_page = pdf.pages[page_num]
        table = current_page.extract_table()

        # Convert the table to a pandas DataFrame
        df = pd.DataFrame(table[1:], columns=table[0])

        # Concatenate the current table with the combined DataFrame
        combined_df = pd.concat([combined_df, df], ignore_index=True)

# Open the Excel file using openpyxl
wb = Workbook()
ws = wb.active

# Set sheet name to file_pdf_name
ws.title = file_pdf_name

# Convert the DataFrame to a list of lists
data = [combined_df.columns.tolist()] + combined_df.values.tolist()

# Write the data to the Excel file
for row in data:
    ws.append(row)

# Auto-fit columns
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Auto-fit rows
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)
        if cell.value is not None and ws.row_dimensions[row[0].row].height is not None:
            ws.row_dimensions[row[0].row].height = 1.2 * ws.row_dimensions[row[0].row].height

# Save the modified Excel file
wb.save(pdf_output_file_path_xlsx)
